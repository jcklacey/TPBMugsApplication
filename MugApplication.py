import time
import os
import urllib.request
import urllib.error
import tkinter as tk
from tkinter import ttk
from pathlib import Path
import shutil
from datetime import datetime
import win32print
import win32ui
from PIL import Image, ImageWin, ImageOps
import sys
import subprocess
from openpyxl import load_workbook
from pdf2image import convert_from_bytes
import requests
import base64


# --- Ensure Pillow is installed ---
try:
    from PIL import Image
except ImportError:
    print("Pillow not found, installing...")
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pillow"])
    from PIL import Image  # try again after installing
from PIL import ImageTk 
import qrcode

#### --- ROOT WINDOW SETUP --- ####
root = tk.Tk()
root.title('TPB MUGS APP')
root.geometry('360x513')
root['bg'] = '#f0f0f0'
root.resizable(False, False)
root.attributes('-topmost', True)

#### --- INTIIAL VARIABLE DECLERATIONS
hotFolderDir = "C:/Users/jackl/OneDrive/Desktop/TPB/TPBMugsApplication/HotFolder/"
os.makedirs(hotFolderDir, exist_ok=True)  # create it if it doesn't exist
fileDumpDir = "C:/Users/jackl/OneDrive/Desktop/TPB/TPBMugsApplication/FileDump/"
os.makedirs(fileDumpDir, exist_ok=True)  # create it if it doesn't exist
mugApplicationDir = "C:/Users/jackl/OneDrive/Desktop/TPB/TPBMugsApplication/"
 
### --- VISUALS
#TPB LOGO
tpb_image_path = mugApplicationDir + "TPB_logo_f0f0f0.png"
tpb_original_image = Image.open(tpb_image_path)
tpb_banner_image = ImageTk.PhotoImage(tpb_original_image) 

#FILERIPPER LOGO
FR_image_path = mugApplicationDir + "fileRipperLogo2.5cm.png"
FR_original_image = Image.open(FR_image_path)
FR_banner_image = ImageTk.PhotoImage(FR_original_image) 

#SUCCESS GIF
gif = Image.open(mugApplicationDir + "ripper.gif")
frames = []
try:
    while True:
        frames.append(ImageTk.PhotoImage(gif.copy()))
        gif.seek(len(frames))  # move to next frame
except EOFError:
    pass

#FAIL ICON
failIcon = Image.open(mugApplicationDir + "fail_icon.png")
failIcon_tk = ImageTk.PhotoImage(failIcon)

#### --- STYLE SET UP
# BUTTONS
style = ttk.Style()
style.theme_use("clam") 
style.configure("My.TButton",
                foreground="black",
                background="yellow",
                font=("Arial", 12, "bold"))
style.map("My.TButton",
          background=[("active", "light blue"), ("!active", "darkgrey")])

style = ttk.Style()
style.theme_use("clam") 
style.configure("My.TFrame",
                background="#f0f0f0",
                bordercolor="black",
                borderwidth=2,
                relief="solid")

# LABELS
style = ttk.Style()
style.theme_use("clam")
style.configure("My.TLabel",
                foreground="black",
                background="#f0f0f0",
                font=("Arial", 10, "bold"))

style = ttk.Style()
style.theme_use("clam")
style.configure("Cursive.TLabel",
                foreground="black",
                background="#f0f0f0",
                font=("Arial", 10, "italic"))

style = ttk.Style()
style.theme_use("clam")
style.configure("HyperLink.TLabel",
                foreground="#196dff",
                background="#f0f0f0",
                font=("Arial", 10, "underline"))

style = ttk.Style()
style.theme_use("clam")
style.configure("HyperLink_Hover.TLabel",
                foreground="#6093ec",
                background="#f0f0f0",
                font=("Arial", 10, "underline"))


### - FUNCTIONS
def toggle_on():
    if errorState==False:
        ripperGIF_label.pack()
        sucess_label.pack()
    if errorState==True:
        failIcon_label.pack()
        fail_label.pack()

def toggle_off():
    ripperGIF_label.pack_forget()     
    sucess_label.pack_forget()   
    failIcon_label.pack_forget()
    fail_label.pack_forget()

def flatten_to_rgb(img, bg_colour =(255,255,255)):
    # Force any image into RGB with a solid background - forcing transparent pixels to white
    #ensure that we have an ALPHA channel
    img = img.convert("RGBA") 
    #Create a white bg
    bg = Image.new("RGB", img.size, bg_colour)
    #iterate over all pixels and replace transparent ones with bg_colour
    alpha = img.getchannel("A")
    # Put image over background using alpha as mask
    bg.paste(img, mask=alpha)

    #Force fully transparent pizels to bg_colour
    pixels = bg.load()
    for y in range(bg.height):
        for x in range (bg.width):
            if alpha.getpixel((x,y)) == 0:
                pixels[x,y] = bg_colour
    return bg

def pdf_to_png(url, output_dir):
    try:
        # Download the PDF file
        response = requests.get(url)
        response.raise_for_status()
        print(f"✅ Downloaded PDF from {url}")
        # Convert to images (usually one per page)
        images = convert_from_bytes(response.content, dpi=300)
        # Save the first page as PNG
        output_path = os.path.join(output_dir, shipmentNo_var.get()+".png")
        images[0].save(output_path, "PNG")
        print(f"✅ Saved PNG preview at {output_path}")
        return output_path

    except Exception as e:
        print(f"❌ Failed to convert PDF: {e}")
        return None

def urlValidityChecker(url):
    """
    Checks the URL and returns a local PNG path.
    For PDFs, converts first page to PNG.
    For image URLs, downloads the image.
    Returns: str (path to PNG) or False on failure.
    """
    global errorState
    
    if not url or str(url).strip() == "":
        fail_label.config(text="imgLink cannot be blank!")
        errorState = True
        return False

    try:
        if url.lower().endswith(".pdf"):
            # Convert PDF to PNG
            output_path = pdf_to_png(url, hotFolderDir)
            if output_path:
                errorState = False
                return output_path
            else:
                fail_label.config(text="PDF could not be converted.")
                errorState = True
                return False
        else:
            # Validate regular image URL
            parsed = urllib.parse.urlparse(url)
            if not parsed.scheme or not parsed.netloc:
                fail_label.config(text="imgLink is not a valid URL!")
                errorState = True
                return False

            req = requests.get(url, stream=True)
            if req.status_code == 200 and "image" in req.headers.get("Content-Type", ""):
                # Save image locally
                filename = f"{shipmentNo_var.get()}.png"
                output_path = os.path.join(hotFolderDir, filename)
                with open(output_path, "wb") as f:
                    f.write(req.content)
                print(f"✅ Saved image from URL at {output_path}")
                errorState = False
                return output_path
            else:
                fail_label.config(text=f"Invalid content type or status: {req.status_code}")
                errorState = True
                return False

    except Exception as e:
        print(f"❌ URL check failed: {e}")
        fail_label.config(text="Invalid or unreachable URL.")
        errorState = True
        return False


def imgLink_webscrape():
    """
    Downloads image or converts PDF for the given imgLink_var.
    Prepares prodfile_copy for further processing.
    """
    global errorState, prodfile, prodfile_copy, prodfile_path

    img_url = str(imgLink_var.get()).strip()
    save_path = urlValidityChecker(img_url)

    if save_path:
        # Open the PNG safely
        try:
            with Image.open(save_path) as prodfile:
                prodfile_copy = flatten_to_rgb(prodfile)
            prodfile_path = save_path
            print(f"✅ Prod image ready at: {prodfile_path}")
            errorState = False
        except Exception as e:
            print(f"❌ Failed to open image: {e}")
            fail_label.config(text="Failed to open PNG.")
            errorState = True
    else:
        fail_label.config(text="URL check failed, cannot fetch image.")
        errorState = True


def qrCode_generate():
    fileCount = 0
    global qrCodeFile
    while True:
        if fileCount == 0:
            filename = str(shipmentNo_var.get()) + "_QRcode.png"
        else:
            filename = str(shipmentNo_var.get()) + "_QRcode-" + str(fileCount) + ".png"
        save_path = os.path.join(hotFolderDir, filename)
        global qrCodeFile_Path
        qrCodeFile_Path = str(save_path)
        if not os.path.exists(save_path):
            break
        fileCount += 1
    img = qrcode.make(str(qrCode_var.get()))
    # Convert the image from RGBA to RGB to replace the transparency with a white BG
    qrCodeFile = flatten_to_rgb(img)
    type(img)
    img.save(save_path)
 
def generateButton():
    global errorState
    global fileCount
    fileCount = 0
    toggle_off()
    imgLink_webscrape()
    if errorState==False:
        qrCode_generate()
        sucess_label.config(text = "Print file pulled... generating Prod file now")
        toggle_on()
        # STEP 1: Create a white background template
        template_width = 2539
        template_height = 1032
        template_bg = Image.new("RGB", (template_width, template_height), (255, 255, 255))
        # STEP 2: Paste prodfile_copy onto the white background, centered
        x_offset = (template_width - prodfile_copy.width) // 2
        y_offset = (template_height - prodfile_copy.height) // 2
        template_bg.paste(prodfile_copy, (x_offset, y_offset))
        # 🔹 At this point, template_bg is your prodfile centered on white
        # STEP 3: Create new image to hold template_bg + QR code
        combined_width = template_bg.width + qrCodeFile.width + 10  # 10px margin between
        combined_height = max(template_bg.height, qrCodeFile.height)
        combined_image = Image.new("RGB", (combined_width, combined_height), (255, 255, 255))
        # STEP 4: Paste the white-backgrounded prodfile first
        combined_image.paste(template_bg, (0, 0))
        # STEP 5: Paste the QR code to the right, vertically centered
        qr_offset_x = template_bg.width + 10  # after prodfile + margin
        qr_offset_y = combined_height - qrCodeFile.height
        combined_image.paste(qrCodeFile, (qr_offset_x, qr_offset_y))
        # STEP 6: Save the result
        filename = f"{shipmentNo_var.get()}-{fileCount}.png"
        save_path = os.path.join(hotFolderDir, filename)
        combined_image.save(save_path)
        print(f"Saved combined image at: {save_path}")
        # STEP 7: Rotate image for your PSD layout
        rotated_image = combined_image.transpose(Image.ROTATE_90)
        rotated_image.save(save_path)
        print("Rotated and saved")
        # STEP 8: Clean up
        prodfile_copy.close()
        qrCodeFile.close()
        combined_image.close()
        template_bg.close()
        os.remove(prodfile_path)
        os.remove(qrCodeFile_Path)
        toggle_off()
        sucess_label.config(text = "Nice! Prod file now in HotFolder")
        toggle_on()
        print("✅ Finished generateButton()")
    if errorState==True: 
        toggle_on()
        return

def resizeToTemplate():
 # -----------------------------------------------
        # Combine our prodfile with a correctly sized template to avoid negative space issues in Photoshop
        # Establish the correct height and width for what file should be
        global template_width
        global template_height
        template_width = 3159 #pixels ----> this is wrong, but will work for now
        template_height = 1125 #pixels ----> this is wrong, but will work for now
        #Create a new template file to paste our prodfile on to
        global prodFileTemplate_image
        prodFileTemplate_image = Image.new("RGB", (template_width, template_height), (255, 255, 255))
        # Establish values to determine where the middle is of our template
        x_offset = (template_width - prodfile_copy.width) // 2
        y_offset = (template_height - prodfile_copy.height) // 2
        # Paste prodfile copy ontop of the template, centered
        prodFileTemplate_image.paste(prodfile_copy, (x_offset, y_offset))
        # Save this new combined image into the specified save path
        filename = str(shipmentNo_var.get()) + "-" + str(fileCount) + ".png"
        print("state check 1")
        save_path = os.path.join(hotFolderDir, filename)
        prodFileTemplate_image.save(str(save_path))
        print("saved in" + str(filename))
        # Close image objects
        prodfile.close()
        #prodFileTemplate_image.close()
        print("state check 2")
        # -----------------------------------------------

def update(ind=0):
    frame = frames[ind]
    ripperGIF_label.configure(image=frame)
    ind = (ind + 1) % len(frames)
    root.after(100, update, ind)     

def fileListGenerator():
    global errorState
    toggle_off()
    # Declare the file paths we'll need
    FileList_scrDir = mugApplicationDir + "FileList.txt"
    FileList_destDir = fileDumpDir + "FileList.txt"
    # Create a list out of all the files in the directory
    prodfilesList = [x for x in os.listdir(hotFolderDir) if x.endswith('.png')]
    prodfilesList.sort()
    print("Length of prodfilesList = " + str(len(prodfilesList)))
    # If there werent any PNG's - report this to user
    if (len(prodfilesList)) == 0:
        fail_label.config(text = "No PNG's in HotFolder to List!")
        errorState = True
        toggle_on()
    else:
        # Check if theres an odd number of files, and if there is, tyhen add a placeholder image to make it even. 
        addPlaceholderImage()
        prodfilesList = [x for x in os.listdir(hotFolderDir) if x.endswith('.png')]
        prodfilesList.sort()
        # Create FileList.txt and write the list of files in pairs
        with open(FileList_scrDir, "w") as f:  
            print('File1\tFile2', file=f)
            for i in range(0, len(prodfilesList), 2):
                if i + 1 < len(prodfilesList):
                    f.write(f"{prodfilesList[i]}\t{prodfilesList[i+1]}\n")
                else:
                    f.write(f"{prodfilesList[i]}\n")
        # Move the FileList.txt from the MugApplication folder to the FileDump folder
        shutil.move(FileList_scrDir, FileList_destDir)
        # Move all files from HotFolder to FileDump
        for filename in os.listdir(hotFolderDir):
            src = os.path.join(hotFolderDir, filename)
            dst = os.path.join(fileDumpDir, filename)
            if os.path.isfile(src):
                shutil.move(src, dst)
        #Display success GIF
        errorState = False
        sucess_label.config(text = "FileList created - files moved to FileDump")
        toggle_on()

def PrintButton():
    global errorState
    #reset success/fail icon
    toggle_off()
    # Convert PSDs to PNGs
    # Make sure PNGs folder exists
    os.makedirs(hotFolderDir, exist_ok=True)
    for filename in os.listdir(hotFolderDir):
        if filename.lower().endswith(".psd"):
            psd_path = os.path.join(hotFolderDir, filename)
            png_filename = os.path.splitext(filename)[0] + ".png"
            png_path = os.path.join(hotFolderDir, png_filename)
            try:
            # Open PSD and convert
                with Image.open(psd_path) as img:
                    dpi = img.info.get("dpi", (300, 300))
                    img.save(png_path, "PNG", dpi=dpi)
                    time.sleep(0.5)
                    print(f"Converted: {filename} → {png_filename}")
                    errorState = False
                    toggle_on()
            except Exception as e:
                print(f"Failed to convert {filename}: {e}") 
                errorState= True
                toggle_on()
            # delete the PSD after you've created the PNG
            os.remove(psd_path) 
    # Delete loose files (excluding folders) in FileDump
    for file in os.listdir(fileDumpDir):
        file_path = os.path.join(fileDumpDir, file)
        if os.path.isfile(file_path):  # only delete files, not folders
            os.remove(file_path)
            print(f"Deleted: {file_path}")
    # Print the PNGs in the HotFolder folder 
    for filename in os.listdir(hotFolderDir):
        if filename.lower().endswith(".png"):
            filepath = os.path.join(hotFolderDir, filename)
            if os.path.isfile(filepath):
                try:
                    # print the file on the computers default printer
                    print(f"Sending {filename} to printer...")
                     # --- Start direct printing ---
                    printer_name = win32print.GetDefaultPrinter()
                    # Create device context
                    dc = win32ui.CreateDC()           # no arguments here
                    dc.CreatePrinterDC(printer_name)  # now pass the printer name
                    # Open the image
                    img = Image.open(filepath)
                    #---START DPI HANDLING------------------
                    dpi = img.info.get("dpi", (300, 300))
                    # Printer DPI (horizontal, vertical)
                    printer_dpi_x = dc.GetDeviceCaps(88)  # LOGPIXELSX
                    printer_dpi_y = dc.GetDeviceCaps(90)  # LOGPIXELSY
                    # Convert image size from pixels → inches
                    width_inch = img.width / dpi[0]
                    height_inch = img.height / dpi[1]
                    # Convert inches → printer units
                    width_printerUnits = int(width_inch * printer_dpi_x)
                    height_prinerUnits = int(height_inch * printer_dpi_y)
                    #---END DPI HANDLING------------------
                    # Start printing
                    dc.StartDoc(filepath)
                    dc.StartPage()
                    # Draw image at correct physical size
                    dib = ImageWin.Dib(img)
                    dib.draw(dc.GetHandleOutput(), (0, 0, width_printerUnits, height_prinerUnits))
                    # Finish printing
                    dc.EndPage()
                    dc.EndDoc()
                    dc.DeleteDC()
                    # --- End direct printing ---
                except Exception as e:
                    print(f"Failed to print {filename}: {e}")
                    errorState=True
                    toggle_on()
                shutil.move(filepath, fileDumpDir)
                errorState=False
                toggle_on() 
    toggle_off() 
    sucess_label.config(text ="Nice! PSD's sent to printer")
    toggle_on() 
    # Create a new archive folder, including todays date
    todaysDate = datetime.today().strftime("%Y-%m-%d")
    newArchiveFolder = f"{todaysDate}_ARCHIVE"
    newArchiveFolder_path = os.path.join(fileDumpDir, newArchiveFolder)
    os.makedirs(newArchiveFolder_path, exist_ok=True)
    # Put all loose files (not folders) in FileDump into the new archive folder
    for filename in os.listdir(fileDumpDir):
        filepath = os.path.join(fileDumpDir, filename)
        if filename.lower().endswith((".png", ".txt")):
            if os.path.isfile(filepath):
                try:
                    shutil.move(filepath, newArchiveFolder_path)
                    # Show Success message
                    errorState=True
                    toggle_on()
                except:
                    print(f"Failed to move {filename}: {e}")
                    # Show Fail message
                    errorState=False
                    toggle_on()

def addPlaceholderImage():
    # Create a list out of all the files in the directory
    prodfilesCount = [x for x in os.listdir(hotFolderDir) if x.endswith('.png')]
    if len(prodfilesCount) % 2 != 0:
        print("prodfilesCount list is odd!")
        placeholderImage = "placeholderImage.png"
        src = os.path.join(mugApplicationDir, placeholderImage)
        dst = os.path.join(hotFolderDir, placeholderImage)
        shutil.copy(src, dst)
        print("placeholerImage moved ocer successfully")
    else:
        print("prodfilesCount list is even!")

def imgLink_webscrape_XLSX_TPB(sheet):
    global prodfile_copy
    global prodfile_path
    global errorState
    global qrCodeFile
    global qrCodeFile_Path
    global fileCount
    for row in sheet.iter_rows(min_row=2):
        if row[0].value and row[2].value and row[3].value:
            # --- Correlate cell values to required variables
            orderNumber = row[0].value
            sku = row[1].value
            try:
                quantity = int(row[2].value)
            except (TypeError, ValueError):
                print(f"Skipping bad quantity in row: {row[2].value}")
                continue
            imgLink = row[3].value
            # ---- BATCH imgLink_WebScraper
            for i in range(quantity):
                fileCount = 0
                while True:
                    filename = (
                        f"{orderNumber}.png" if fileCount == 0
                        else f"{orderNumber}_{fileCount}.png")
                    save_path = os.path.join(hotFolderDir, filename)
                    if not os.path.exists(save_path):
                        break
                    fileCount += 1
                if urlValidityChecker(imgLink):
                    urllib.request.urlretrieve(imgLink, save_path)  
                    with Image.open(save_path) as prodfile:
                        prodfile_copy = flatten_to_rgb(prodfile)
                    prodfile_path = save_path
                    print("✅ Prod image downloaded:", prodfile_path)
                    errorState = False
                else:
                    print("❌ Invalid image link:", imgLink)
                    errorState = True
                # ---- BATCH QR Code generator
                qr_fileCount = 0
                global qrCodeFile
                while True:
                    if qr_fileCount == 0:
                        filename = str(orderNumber) + "_QRcode.png"
                    else:
                        filename = str(orderNumber) + "_QRcode-" + str(fileCount) + ".png"
                    save_path = os.path.join(hotFolderDir, filename)
                    global qrCodeFile_Path
                    qrCodeFile_Path = str(save_path)
                    if not os.path.exists(save_path):
                        break
                    qr_fileCount += 1
                img = qrcode.make(orderNumber)
                qrCodeFile = flatten_to_rgb(img)
                img.save(save_path)
                # ✅ generate combined prod + QR file with unique name
                filename_combined = f"{orderNumber}_{i}.png"
                generate_XLSX(filename_combined)

def imgLink_webscrape_XLSX_FE(sheet):
    global prodfile_copy
    global prodfile_path
    global errorState
    global qrCodeFile
    global qrCodeFile_Path
    global fileCount
    for row in sheet.iter_rows(min_row=2):
        if row[0].value and row[2].value and row[3].value:
            # --- Correlate cell values to required variables
            orderNumber = row[0].value
            sku = row[1].value
            try:
                quantity = int(row[2].value)
            except (TypeError, ValueError):
                print(f"Skipping bad quantity in row: {row[2].value}")
                continue
            imgLink = row[3].value
            # ---- BATCH imgLink_WebScraper
            for i in range(quantity):
                fileCount = 0
                while True:
                    filename = (
                        f"{orderNumber}.png" if fileCount == 0
                        else f"{orderNumber}_{fileCount}.png")
                    save_path = os.path.join(hotFolderDir, filename)
                    if not os.path.exists(save_path):
                        break
                    fileCount += 1
                if urlValidityChecker(imgLink):
                    urllib.request.urlretrieve(imgLink, save_path)  
                    with Image.open(save_path) as prodfile:
                        prodfile_copy = flatten_to_rgb(prodfile)
                    prodfile_path = save_path
                    print("✅ Prod image downloaded:", prodfile_path)
                    errorState = False
                else:
                    print("❌ Invalid image link:", imgLink)
                    errorState = True
                # ---- BATCH QR Code generator
                qr_fileCount = 0
                global qrCodeFile
                while True:
                    if qr_fileCount == 0:
                        filename = str(orderNumber) + "_QRcode.png"
                    else:
                        filename = str(orderNumber) + "_QRcode-" + str(fileCount) + ".png"
                    save_path = os.path.join(hotFolderDir, filename)
                    global qrCodeFile_Path
                    qrCodeFile_Path = str(save_path)
                    if not os.path.exists(save_path):
                        break
                    qr_fileCount += 1
                img = qrcode.make(orderNumber)
                qrCodeFile = flatten_to_rgb(img)
                img.save(save_path)
                # ✅ generate combined prod + QR file with unique name
                filename_combined = f"{orderNumber}_{i}.png"
                generate_XLSX(filename_combined)

def imgLink_webscrape_XLSX_FE(sheet):
    global prodfile_copy
    global prodfile_path
    global errorState
    global qrCodeFile
    global qrCodeFile_Path
    global fileCount
    for row in sheet.iter_rows():
        if row[0].value and row[2].value and row[3].value:
            # --- Correlate cell values to required variables
            qrCode = row[0].value
            colourVariation = row[7].value
            try:
                quantity = int(row[15].value)
            except (TypeError, ValueError):
                print(f"Skipping bad quantity in row: {row[0].value}")
                continue
            imgLink = row[23].value
            # ---- BATCH imgLink_WebScraper
            for i in range(quantity):
                fileCount = 0
                while True:
                    filename = (
                        f"{qrCode}.png" if fileCount == 0
                        else f"{qrCode}_{fileCount}.png")
                    save_path = os.path.join(hotFolderDir, filename)
                    if not os.path.exists(save_path):
                        break
                    fileCount += 1
                if urlValidityChecker(imgLink):
                    urllib.request.urlretrieve(imgLink, save_path)  
                    with Image.open(save_path) as prodfile:
                        prodfile_copy = flatten_to_rgb(prodfile)
                    prodfile_path = save_path
                    print("✅ Prod image downloaded:", prodfile_path)
                    errorState = False
                else:
                    print("❌ Invalid image link:", imgLink)
                    errorState = True
                # ---- BATCH QR Code generator
                qr_fileCount = 0
                global qrCodeFile
                while True:
                    if qr_fileCount == 0:
                        filename = str(qrCode) + "_QRcode.png"
                    else:
                        filename = str(qrCode) + "_QRcode-" + str(fileCount) + ".png"
                    save_path = os.path.join(hotFolderDir, filename)
                    global qrCodeFile_Path
                    qrCodeFile_Path = str(save_path)
                    if not os.path.exists(save_path):
                        break
                    qr_fileCount += 1
                img = qrcode.make(qrCode)
                img = img.convert("RGB")
                if colourVariation == "Black":
                    border_size = 10
                    img = ImageOps.expand(img, border=border_size, fill="black")
                qrCodeFile = flatten_to_rgb(img)
                img.save(save_path)
                # ✅ generate combined prod + QR file with unique name
                filename_combined = f"{qrCode}_{i}.png"
                generate_XLSX(filename_combined)

def generate_XLSX(filename):
    global prodfile_copy
    # STEP 1: Create a white background template
    template_width = 2539
    template_height = 1032
    template_bg = Image.new("RGB", (template_width, template_height), (255, 255, 255))
    # STEP 2: Paste prodfile_copy onto the white background, centered
    x_offset = (template_width - prodfile_copy.width) // 2
    y_offset = (template_height - prodfile_copy.height) // 2
    template_bg.paste(prodfile_copy, (x_offset, y_offset))
    # 🔹 At this point, template_bg is your prodfile centered on white
    # STEP 3: Create new image to hold template_bg + QR code
    combined_width = template_bg.width + qrCodeFile.width + 10  # 10px margin between
    combined_height = max(template_bg.height, qrCodeFile.height)
    combined_image = Image.new("RGB", (combined_width, combined_height), (255, 255, 255))
    # STEP 4: Paste the white-backgrounded prodfile first
    combined_image.paste(template_bg, (0, 0))
    # STEP 5: Paste the QR code to the right, vertically centered
    qr_offset_x = template_bg.width + 10  # after prodfile + margin
    qr_offset_y = combined_height - qrCodeFile.height
    combined_image.paste(qrCodeFile, (qr_offset_x, qr_offset_y))
    # STEP 6: Save the result
    save_path = os.path.join(hotFolderDir, filename)
    combined_image.save(save_path)
    print(f"Saved combined image at: {save_path}")
    # STEP 7: Rotate image for your PSD layout
    rotated_image = combined_image.transpose(Image.ROTATE_90)
    rotated_image.save(save_path)
    print("Rotated and saved")
    # STEP 8: Clean up
    prodfile_copy.close()
    qrCodeFile.close()
    combined_image.close()
    template_bg.close()
    os.remove(prodfile_path)
    os.remove(qrCodeFile_Path)
    print("generate_XLSX() acknowledged")

def xlsxUpload_click(event):
    global prodfile_copy
    global errorState
    print("Label clicked – starting upload…")  # debug
    for file in os.listdir(hotFolderDir):
        if file.endswith(".xlsx") and file.startswith("TPB"):
            filepath = os.path.join(hotFolderDir, file)
            workbook = load_workbook(filepath)
            sheet = workbook.active
            imgLink_webscrape_XLSX_TPB(sheet)
            
        if file.endswith(".xlsx") and file.startswith("FE"):
            filepath = os.path.join(hotFolderDir, file)
            workbook = load_workbook(filepath)
            sheet = workbook.active
            imgLink_webscrape_XLSX_FE(sheet)
    else:
        print("No .xlsx file in Hotfolder")

def xlsxUpload_enter(event):
    xlsxUpload_label.config(style="HyperLink_Hover.TLabel")
    
def xlsxUpload_leave(event):
    xlsxUpload_label.config(style="HyperLink.TLabel")
    
#### --- WINDOW ELEMENT PROPERTIES --- ####
# FRAME
mainFrame = ttk.Frame(root, style='My.TFrame')
mainFrame.pack(padx=20, pady=20, fill="both", expand=True)
# mainWindow LABEL
# mainWindow_label = ttk.Label(mainFrame, text='FILE RIPPER', background="grey", font=("Arial", 16,"bold"))
# mainWindow_label.pack(pady=20)
# FE banner LABEL
FR_banner_label = tk.Label(mainFrame, image=FR_banner_image, border=0)
FR_banner_label.pack(pady=3)
# TPB banner LABEL
tpb_banner_label = tk.Label(mainFrame, image=tpb_banner_image, border=0)
tpb_banner_label.pack()

#----
# imgLink LABEL
imgLink_label = ttk.Label(mainFrame, text='imgLink:', style='My.TLabel')
imgLink_label.pack()
# imgLink ENTRY
imgLink_var = tk.StringVar()
imgLink_entry = ttk.Entry(mainFrame, textvariable=imgLink_var)
imgLink_entry.pack(pady=5)
#----
# qrCode LABEL
qrCode_label = ttk.Label(mainFrame, text='QR Code:', style='My.TLabel')
qrCode_label.pack()
# qrCode ENTRY
qrCode_var = tk.StringVar()
qrCode_entry = ttk.Entry(mainFrame, textvariable=qrCode_var)
qrCode_entry.pack(pady=5)
#----
# shipmentNo LABEL
shipmentNo_label = ttk.Label(mainFrame, text='Shipment Number:', style='My.TLabel')
shipmentNo_label.pack()
# shipmentNo ENTRY
shipmentNo_var = tk.StringVar()
shipmentNo_entry = ttk.Entry(mainFrame, textvariable=shipmentNo_var)
shipmentNo_entry.pack(pady=5)
#----
# .xlsx upload LABEL
xlsxUpload_label = ttk.Label(mainFrame, text='.xlsx upload', style='HyperLink.TLabel')
xlsxUpload_label.pack()
xlsxUpload_label.bind("<Button-1>", xlsxUpload_click)
xlsxUpload_label.bind("<Enter>", xlsxUpload_enter)
xlsxUpload_label.bind("<Leave>", xlsxUpload_leave)
#----
# Generate BUTTON
Generate_button = ttk.Button(mainFrame, text='1.PULL FILE', style='My.TButton', width=20 , command=generateButton)
Generate_button.pack(pady=3)
#----
# fileListGenerate BUTTON
fileListGenerate_button = ttk.Button(mainFrame, text='2.CREATE LIST', style='My.TButton', width=20, command=fileListGenerator)
fileListGenerate_button.pack(pady=3)
#----
# Print BUTTON
Print_button = ttk.Button(mainFrame, text='3.PRINT', style='My.TButton' , width=20, command=PrintButton)
Print_button.pack(pady=3)
#----

# Ripper GIF LABEL
ripperGIF_label = tk.Label(mainFrame, border=0)
# success LABEL
sucess_label = ttk.Label(mainFrame, text='nice!', style='Cursive.TLabel')
# Fail_icon LABEL
failIcon_label = ttk.Label(mainFrame, border=0, image=failIcon_tk)
# fail LABEL
fail_label = ttk.Label(mainFrame, text='Uh Oh - no good!', style='Cursive.TLabel')

# Has to be at the very end of the program. 
update()
root.mainloop()